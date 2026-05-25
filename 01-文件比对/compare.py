import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MERGE_KEY = "区域"
LOC_COLS = ["省份", "市", "区县"]
COMPARE_COLS = ["并网箱", "逆变器"]

RED_FONT = Font(color="FF0000")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def val_changed(a: object, b: object) -> bool:
    a_str = "" if pd.isna(a) else str(a).strip()
    b_str = "" if pd.isna(b) else str(b).strip()
    return a_str != b_str


def change_type(row: pd.Series) -> str:
    parts = []
    if row["并网箱_变化"]:
        parts.append("并网箱变化")
    if row["逆变器_变化"]:
        parts.append("逆变器变化")
    return " + ".join(parts)


def _auto_width(ws: Workbook, headers: list[str]) -> None:
    """根据表头和内容自适应列宽"""
    for col_idx in range(1, len(headers) + 1):
        # 以表头长度为基准
        max_len = len(str(headers[col_idx - 1]))
        # 检查前 50 行内容，取最大宽度
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 51),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                # 中文字符约按 2 倍宽度计算
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_len:
                    max_len = length
        # 设置列宽，加上少量余量
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)


def _style_header(ws: Workbook) -> None:
    """给表头行应用配色和边框"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def write_sheet1(wb: Workbook, diff: pd.DataFrame) -> None:
    """变更明细：旧值/新值，未变化留空，新值标红"""
    ws = wb.active
    ws.title = "变更明细"

    headers = LOC_COLS + [MERGE_KEY]
    for col in COMPARE_COLS:
        headers += [f"{col}_旧", f"{col}_新"]
    headers.append("变化类型")

    ws.append(headers)

    for _, row in diff.iterrows():
        values = [row[c] for c in LOC_COLS] + [row[MERGE_KEY]]
        red_cells: list[int] = []

        for i, col in enumerate(COMPARE_COLS):
            old_col = f"{col}_旧"
            new_col = f"{col}_新"
            changed = row[f"{col}_变化"]
            if changed:
                values.append(row[old_col])
                values.append(row[new_col])
                # 新值列在行中的索引
                new_col_idx = len(LOC_COLS) + 1 + i * 2 + 1
                red_cells.append(new_col_idx)
            else:
                values.append("")
                values.append("")

        values.append(row["变化类型"])

        row_idx = ws.max_row + 1
        ws.append(values)

        # 数据单元格样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        for col_idx in red_cells:
            ws.cell(row=row_idx, column=col_idx + 1).font = RED_FONT

    # 行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22

    _style_header(ws)
    _auto_width(ws, headers)


def write_sheet2(wb: Workbook, diff: pd.DataFrame) -> None:
    """最新数据：去掉旧值列，并网箱/逆变器展示最新值"""
    ws = wb.create_sheet("最新数据")

    headers = LOC_COLS + [MERGE_KEY] + COMPARE_COLS + ["变化类型"]
    ws.append(headers)

    for _, row in diff.iterrows():
        values = [row[c] for c in LOC_COLS] + [row[MERGE_KEY]]
        for col in COMPARE_COLS:
            values.append(row[f"{col}_新"])
        values.append(row["变化类型"])

        row_idx = ws.max_row + 1
        ws.append(values)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22

    _style_header(ws)
    _auto_width(ws, headers)


def run_compare(old_file: str, new_file: str, output_file: str) -> int:
    """执行对比并输出 Excel，返回变更记录数"""
    df_old = pd.read_excel(old_file)
    df_new = pd.read_excel(new_file)

    required_old = LOC_COLS + [MERGE_KEY] + COMPARE_COLS
    required_new = [MERGE_KEY] + COMPARE_COLS + LOC_COLS
    for cols, df, name in [(required_old, df_old, old_file), (required_new, df_new, new_file)]:
        missing = [c for c in cols if c not in df.columns]
        if missing:
            raise ValueError(f"{name} 缺少列: {missing}")

    merged = pd.merge(
        df_old[required_old].rename(columns={c: f"{c}_旧" for c in LOC_COLS}),
        df_new[required_new].rename(columns={c: f"{c}_新" for c in LOC_COLS}),
        on=MERGE_KEY,
        how="outer",
        suffixes=("_旧", "_新"),
    )

    for col in LOC_COLS:
        merged[col] = merged[f"{col}_旧"].combine_first(merged[f"{col}_新"])
        merged.drop(columns=[f"{col}_旧", f"{col}_新"], inplace=True)

    for col in COMPARE_COLS:
        merged[f"{col}_变化"] = merged.apply(
            lambda r, c=col: val_changed(r[f"{c}_旧"], r[f"{c}_新"]), axis=1
        )

    changed_cols = [f"{col}_变化" for col in COMPARE_COLS]
    diff = merged[merged[changed_cols].any(axis=1)].copy()
    diff["变化类型"] = diff.apply(change_type, axis=1)
    diff = diff.reset_index(drop=True)

    wb = Workbook()
    write_sheet1(wb, diff)
    write_sheet2(wb, diff)
    wb.save(output_file)

    return len(diff)
